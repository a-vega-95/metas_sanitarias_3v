import sys
import os
import csv
import openpyxl

# Add project root to path
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(current_dir))
sys.path.append(os.path.join(project_root, 'src'))

from modules.dataloaders import scan_rem_files, load_piv_for_year, get_rem_sheet_data, load_poblacion_a_cargo
from modules.utils import normalize_path, log_cuarentena_valor_invalido
import config

def to_num(val):
    if val is None: return 0
    if isinstance(val, (int, float)): return val
    try:
        return float(val)
    except:
        return 0

def calcular_meta_7(agno=None, max_mes=None):
    # Parameters
    agno_eval = int(agno) if agno else config.AGNO_ACTUAL
    m_limit = int(max_mes) if max_mes else int(os.environ.get('MAX_MES', 12))
    agno_piv = agno_eval - 1

    print(f"=== Calculando Meta 7: Enfermedades Respiratorias (Asma/EPOC) ({agno_eval}, Mes {m_limit}) ===")
    
    # 3. Cargar parámetros y prevalencias
    params_m7 = config.get_meta_params('Meta_7', agno_eval)
    META_FIJADA_7 = params_m7.get('fijada', 16.77)
    META_NACIONAL_7 = params_m7.get('nacional', 15.0)
    
    PREV_ASMA = config.get_prevalencia('RESP', 'Asma_5_mas', agno_eval)
    PREV_EPOC = config.get_prevalencia('RESP', 'EPOC_40_mas', agno_eval)
    
    # 4. Configuración Coordenadas
    coords = config.get_meta_coordinates('Meta_7', agno_eval)
    SHEET_TARGET = coords.get('hoja', "P3")
    coords_asma = coords.get('asma', {"fila": 65, "col_ini": "H", "col_fin": "AM"})
    coords_epoc = coords.get('epoc', {"fila": 69, "col_ini": "V", "col_fin": "AM"})
    
    # 5. Cargar archivo PIV
    piv_data = load_piv_for_year(agno_piv)
    if not piv_data:
        print(f"[WARNING] No se encontraron datos PIV para {agno_piv}. Los denominadores serán 0.")
        piv_data = [] # Continuar con ceros

    dir_p = config.get_serie_p_corte(m_limit, agno_eval)
    if dir_p is None:
        print(f"[Meta 7] Sin corte Serie P disponible para m_limit={m_limit}. Se reportaran denominadores (PIV) con numeradores en 0.")
    mapping = scan_rem_files(dir_p) if dir_p else []


    # 1. Denominadores Estimados (PIV)
    denominadores = {}
    
    for row in piv_data:
        centro = row.get('COD_CENTRO', '')
        edad = row.get('EDAD_EN_FECHA_CORTE')
        if edad is None: edad = -1
        estado = row.get('ACEPTADO_RECHAZADO', '')
        
        if estado == 'ACEPTADO':
            if centro not in denominadores:
                denominadores[centro] = 0
            
            # Asma 5+
            if edad >= 5:
                denominadores[centro] += (1 * PREV_ASMA)
                
            # EPOC 40+
            if edad >= 40:
                denominadores[centro] += (1 * PREV_EPOC)

                
    # Redondear
    denominadores = {k: round(v) for k, v in denominadores.items()}
    
    # Aplicar Override Manual (Población a Cargo) si el denominador PIV es 0
    poblacion_manual = load_poblacion_a_cargo(agno_eval)
    centros_en_rem = set()
    for entry in mapping:
        c = entry['code']
        if c[-1].isalpha() and c[:-1].isdigit():
            c = c[:-1]
        centros_en_rem.add(c)
        
    for c in (set(denominadores.keys()) | centros_en_rem):
        if denominadores.get(c, 0) == 0 and c in poblacion_manual and 'Meta_7' in poblacion_manual[c]:
            denominadores[c] = poblacion_manual[c]['Meta_7']
    
    # 2. Numeradores (REM P3)
    numeradores = {}
    
    for entry in mapping:
        raw_code = entry['code']
        code = raw_code
        if raw_code[-1].isalpha() and raw_code[:-1].isdigit():
            code = raw_code[:-1]
            
        file_path = entry['path']
            
        if code not in numeradores:
            numeradores[code] = 0
            
        if not os.path.exists(file_path): continue
        
        # Usar la nueva función cacheada
        from openpyxl.utils import get_column_letter, column_index_from_string
        
        # Generar lista de columnas
        cols_asma = [get_column_letter(i) for i in range(column_index_from_string(coords_asma['col_ini']), column_index_from_string(coords_asma['col_fin']) + 1)]
        cols_epoc = [get_column_letter(i) for i in range(column_index_from_string(coords_epoc['col_ini']), column_index_from_string(coords_epoc['col_fin']) + 1)]
        all_cols = sorted(list(set(cols_asma) | set(cols_epoc)))
        rows_needed = sorted(list({coords_asma['fila'], coords_epoc['fila']}))
        
        sheet_data = get_rem_sheet_data(file_path, SHEET_TARGET, rows_needed, all_cols)
        
        # Asma
        for c in cols_asma:
            val = sheet_data.get((coords_asma['fila'], c))
            if val is None or not isinstance(val, (int, float)):
                log_cuarentena_valor_invalido(file_path, SHEET_TARGET, f"{c}{coords_asma['fila']}", val)
            else:
                numeradores[code] += val
                
        # EPOC
        for c in cols_epoc:
            val = sheet_data.get((coords_epoc['fila'], c))
            if val is None or not isinstance(val, (int, float)):
                log_cuarentena_valor_invalido(file_path, SHEET_TARGET, f"{c}{coords_epoc['fila']}", val)
            else:
                numeradores[code] += val
            
    # 3. Agrupación Específica Meta 7 (CESCOF -> CESFAM)
    PARENT_MAP_META_7 = {
        '121788': '121307', # Las Quilas -> Amanecer
        '121780': '121347', # El Salar -> Pedro de Valdivia
        '121782': '121305', # Arquenco -> Villa Alegre
    }
    
    for child, parent in PARENT_MAP_META_7.items():
        # Transferir denominadores
        if child in denominadores:
            if parent not in denominadores:
                denominadores[parent] = 0
            denominadores[parent] += denominadores[child]
            del denominadores[child]
            
        # Transferir numeradores
        if child in numeradores:
            if parent not in numeradores:
                numeradores[parent] = 0
            numeradores[parent] += numeradores[child]
            del numeradores[child]

    # Reporte
    reporte = []
    all_centers = set(denominadores.keys()) | set(numeradores.keys())
    
    total_num = 0
    total_den = 0
    
    for c in all_centers:
        num = numeradores.get(c, 0)
        den = denominadores.get(c, 0)
        cump = (num/den*100) if den > 0 else 0
        
        total_num += num
        total_den += den
        
        params = config.get_meta_params('Meta_7', cod_centro=c, agno=agno_eval)
        reporte.append({
            'Centro': c, 'Meta_ID': 'Meta 7', 'Indicador': 'Resp (Asma/EPOC)',
            'Numerador': num, 'Denominador': den, 'Cumplimiento': cump,
            'Meta_Fijada': params.get('fijada', 16.77), 'Meta_Nacional': params.get('nacional', 15.0)
        })
        
    print("\n=== RESULTADOS GLOBALES META 7 ===")
    print(f"Numerador Total: {total_num}")
    print(f"Denominador Total (Est): {total_den}")
    if total_den > 0:
         print(f"Cumplimiento: {total_num/total_den*100:.2f}%")
         
    return reporte
    
if __name__ == "__main__":
    import config
    res = calcular_meta_7(config.AGNO_ACTUAL, int(os.environ.get('MAX_MES', 12)))
    for r in res: print(r)

