import os
import csv
import openpyxl
import pyarrow.parquet as pq
from modules.utils import normalize_path, log_cuarentena_denegado, log_cuarentena_valor_invalido

# Cache global para evitar re-lectura de archivos durante corridas históricas
_CELL_CACHE = {}  # { (file_path, sheet, cell): value }
_PIV_CACHE = {}   # { year: data_list }
_POBLACION_CARGO_CACHE = {} # { year: { cod_centro: { meta_id: denominador } } }

def extract_date_from_path(file_path):
    """
    Extracts year and month from a file path.
    Returns (year, month_int) or (None, None).
    Month is 1-12.
    """
    path_parts = normalize_path(file_path).split(os.sep)
    
    year = None
    month = None
    
    # Grid of month names to int
    MONTH_MAP = {
        'ENE': 1, 'JAN': 1, 'ENERO': 1, 'JANUARY': 1,
        'FEB': 2, 'FEBRERO': 2, 'FEBRUARY': 2,
        'MAR': 3, 'MARZO': 3, 'MARCH': 3,
        'ABR': 4, 'APR': 4, 'ABRIL': 4, 'APRIL': 4,
        'MAY': 5, 'MAYO': 5,
        'JUN': 6, 'JUNIO': 6, 'JUNE': 6,
        'JUL': 7, 'JULIO': 7, 'JULY': 7,
        'AGO': 8, 'AUG': 8, 'AGOSTO': 8, 'AUGUST': 8,
        'SEP': 9, 'SEPT': 9, 'SEPTIEMBRE': 9, 'SEPTEMBER': 9,
        'OCT': 10, 'OCTUBRE': 10, 'OCTOBER': 10,
        'NOV': 11, 'NOVIEMBRE': 11, 'NOVEMBER': 11,
        'DIC': 12, 'DEC': 12, 'DICIEMBRE': 12, 'DECEMBER': 12
    }
    
    for part in path_parts:
        part_upper = part.upper()
        
        # Check for Year (4 digits starting with 20) - exact match
        if part.isdigit() and len(part) == 4 and part.startswith('20'):
            year = int(part)
        else:
            # Also check for year embedded in folder names like 'REM_A_2025'
            tokens = part.replace('_', ' ').replace('-', ' ').split()
            for token in tokens:
                if token.isdigit() and len(token) == 4 and token.startswith('20'):
                    year = int(token)
            
        # Check for Month as numeric string (01-12) - exact 2-digit folder name
        if part.isdigit() and len(part) == 2:
            val = int(part)
            if 1 <= val <= 12:
                month = val
            
        # Check for Month as text name (ENE, ENERO, etc.)
        subparts = part_upper.replace('_', ' ').replace('-', ' ').split()
        for sub in subparts:
            if sub in MONTH_MAP:
                month = MONTH_MAP[sub]
                
    return year, month

class FileIndexer:
    """Global singleton to scan directories exactly once per session."""
    _cache = {}
    
    @classmethod
    def get_files(cls, dir_path):
        abs_path = normalize_path(dir_path)
        if abs_path not in cls._cache:
            cls._cache[abs_path] = cls._scan(abs_path)
        return cls._cache[abs_path]
        
    @classmethod
    def _scan(cls, abs_path):
        from modules.utils import setup_audit_logger, load_center_names, log_cuarentena_denegado
        logger = setup_audit_logger()
        valid_centers_map = load_center_names()
        
        mapping = []
        if not os.path.exists(abs_path):
            logger.error(f"Directorio de origen no encontrado: {abs_path}")
            return []
            
        logger.info(f"Escaneando disco (1ra y unica vez): {abs_path}")
        
        denegados = 0
        validos = 0
        
        for root, dirs, files in os.walk(abs_path):
            for filename in files:
                if not (filename.lower().endswith('.xlsm') or filename.lower().endswith('.xlsx')):
                    continue

                full_path = os.path.join(root, filename)
                year, month = extract_date_from_path(full_path)
                
                raw_code = os.path.splitext(filename)[0].upper()
                code = raw_code
                if code and code[-1].isalpha() and code[:-1].isdigit():
                    code = code[:-1]
                
                if code not in valid_centers_map and raw_code not in valid_centers_map:
                    log_cuarentena_denegado(full_path, code)
                    denegados += 1
                    continue

                mapping.append({
                    'path': full_path,
                    'year': year,
                    'month': month,
                    'filename': filename,
                    'code': code
                })
                validos += 1
                
        logger.info(f"Escaneo particion {os.path.basename(abs_path)} finalizado | Validos: {validos} | Denegados: {denegados}")
        return mapping

def scan_rem_files(root_dir):
    """Entry point compatible para las metas que usan el indexador."""
    return FileIndexer.get_files(root_dir)

def get_rem_value(file_path, sheet_name, cell_coordinate):
    """
    Retrieves a value from a cell. Uses cache to avoid re-opening files.
    Returns 0 if empty or invalid, logging to Quarantine only on the first failed read.
    """
    if not os.path.exists(file_path):
        return 0
        
    cache_key = (file_path, sheet_name, cell_coordinate)
    if cache_key in _CELL_CACHE:
        return _CELL_CACHE[cache_key]
        
    try:
        from modules.utils import log_cuarentena_valor_invalido
        # read_only is crucial to prevent hangs in complex Excel files
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            _CELL_CACHE[cache_key] = 0
            return 0
            
        sheet = wb[sheet_name]
        val = sheet[cell_coordinate].value
        wb.close()
        
        # Validacion de datos antes de guardar en cache
        if val is None:
            log_cuarentena_valor_invalido(file_path, sheet_name, cell_coordinate, val)
            result = 0
        elif isinstance(val, (int, float)):
            result = val
        else:
            # Texto en celda numerica
            log_cuarentena_valor_invalido(file_path, sheet_name, cell_coordinate, val)
            result = 0
            
        _CELL_CACHE[cache_key] = result
        return result
        
    except Exception as e:
        print(f"[ERROR] Reading {os.path.basename(file_path)} [{sheet_name}!{cell_coordinate}]: {e}")
        _CELL_CACHE[cache_key] = 0
        return 0
def get_rem_sheet_data(file_path, sheet_name, rows_needed, columns_needed):
    """
    Lee multiples celdas de una hoja en una sola pasada y las devuelve como dict.
    Usa cache para evitar re-lecturas.
    rows_needed: list of ints
    columns_needed: list of strings (e.g. ['J', 'K'])
    Returns: { (row, col): value }
    """
    if not os.path.exists(file_path):
        return {}
        
    results = {}
    pending_rows = []
    
    # Check cache first
    for r in rows_needed:
        for c in columns_needed:
            cache_key = (file_path, sheet_name, f"{c}{r}")
            if cache_key in _CELL_CACHE:
                results[(r, c)] = _CELL_CACHE[cache_key]
            else:
                if r not in pending_rows:
                    pending_rows.append(r)
    
    if not pending_rows:
        return results
        
    try:
        from openpyxl.utils import column_index_from_string
        # Open once
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return results
            
        sheet = wb[sheet_name]
        min_r = min(pending_rows)
        max_r = max(pending_rows)
        col_indices = {c: column_index_from_string(c) - 1 for c in columns_needed}
        
        # Iter rows once
        for r_idx, row in enumerate(sheet.iter_rows(min_row=min_r, max_row=max_r), start=min_r):
            if r_idx in pending_rows:
                for col_letter, c_idx in col_indices.items():
                    val = row[c_idx].value
                    
                    # Store in results and cache
                    results[(r_idx, col_letter)] = val
                    _CELL_CACHE[(file_path, sheet_name, f"{col_letter}{r_idx}")] = val
        
        wb.close()
        return results
    except Exception as e:
        print(f"[ERROR] Batch reading {os.path.basename(file_path)}: {e}")
        return results

def load_piv_for_year(year):
    """
    Busca, valida el esquema y carga el archivo PIV correspondiente al año indicado.
    Retorna una lista de diccionarios con cache.
    """
    if year in _PIV_CACHE:
        return _PIV_CACHE[year]

    import pyarrow.parquet as pq
    from config import DATOS_DIR
    
    # 1. Intentar buscar en la subcarpeta del año (Estructura Jerárquica)
    piv_dir_year = normalize_path(os.path.join(DATOS_DIR, "raw", "piv", str(year)))
    piv_dir_base = normalize_path(os.path.join(DATOS_DIR, "raw", "piv"))
    
    piv_files = []
    target_dir = piv_dir_base
    
    if os.path.exists(piv_dir_year):
        found_in_year = [f for f in os.listdir(piv_dir_year) if f.startswith(f"PIV_{year}_") and f.endswith(".parquet")]
        if found_in_year:
            piv_files = found_in_year
            target_dir = piv_dir_year
            
    # 2. Si no se encuentra en la carpeta del año, buscar en la raíz de piv (Compatibilidad)
    if not piv_files:
        if os.path.exists(piv_dir_base):
            piv_files = [f for f in os.listdir(piv_dir_base) if f.startswith(f"PIV_{year}_") and f.endswith(".parquet")]
            if not piv_files:
                # Intento desesperado: buscar cualquier PIV del año
                piv_files = [f for f in os.listdir(piv_dir_base) if f.startswith("PIV_") and f.endswith(".parquet")]
                piv_files = [f for f in piv_files if str(year) in f]
                
                if not piv_files:
                    # Warning silenciado si no hay archivos (se asume 0 poblacion)
                    _PIV_CACHE[year] = []
                    return []
    
    try:
        piv_files.sort(reverse=True)
        target_file = os.path.join(target_dir, piv_files[0])
        
        required_columns = ['COD_CENTRO', 'EDAD_EN_FECHA_CORTE', 'ACEPTADO_RECHAZADO', 'GENERO', 'GENERO_NORMALIZADO']
        
        schema = pq.read_schema(target_file)
        missing = [c for c in required_columns if c not in schema.names]
        if missing:
            print(f"[ERROR] Faltan columnas en PIV {year}: {missing}")
            return []
            
        table = pq.read_table(target_file, columns=required_columns)
        data = table.to_pylist()
        _PIV_CACHE[year] = data
        print(f"[PIV] Cacheado PIV {year} con {len(data)} registros.")
        return data
    except Exception as e:
        print(f"[ERROR] Fallo cargando PIV {year}: {e}")
        _PIV_CACHE[year] = []
        return []

def load_poblacion_a_cargo(year):
    """
    Carga denominadores manuales para centros que no tienen PIV oficial.
    Retorna: { 'COD_CENTRO': { 'Meta_ID': denominador } }
    """
    if year in _POBLACION_CARGO_CACHE:
        return _POBLACION_CARGO_CACHE[year]
        
    from config import DATOS_DIR
    csv_path = normalize_path(os.path.join(DATOS_DIR, "dictionaries", "poblacion_a_cargo.csv"))
    
    overrides = {}
    if not os.path.exists(csv_path):
        _POBLACION_CARGO_CACHE[year] = overrides
        return overrides
        
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    row_year = int(row.get('AGNO', year))
                    if row_year != year:
                        continue
                        
                    cod = row.get('COD_CENTRO', '').strip()
                    meta = row.get('META_ID', '').strip()
                    den = float(row.get('DENOMINADOR', 0))
                    
                    if cod not in overrides:
                        overrides[cod] = {}
                    overrides[cod][meta] = den
                except ValueError:
                    continue
                    
        _POBLACION_CARGO_CACHE[year] = overrides
        return overrides
    except Exception as e:
        print(f"[ERROR] Fallo leyendo poblacion_a_cargo.csv: {e}")
        _POBLACION_CARGO_CACHE[year] = {}
        return {}
